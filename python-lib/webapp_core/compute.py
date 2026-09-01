"""Validate an input file against a category recipe, evaluate the computed columns and
build the output workbook (canonical values + **live Excel formulas** in the computed
columns + lookup sheets + a Parameters sheet).

Pure module: inputs are bytes / dicts, outputs are bytes / DataFrames. No ``dataiku``.
"""

from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from .sample_parser import detect_header_row, _headers_from_row, _norm, col_letter_to_index

_A1_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?(\d+)$")
_TOK_RE = re.compile(r"\{tok:([^}]+)\}")


_SCAN_ROWS = 60


def _nkey(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s)).strip().lower() if s is not None else ""


# --------------------------------------------------------------------------- input IO
def read_table(file_bytes: bytes, filename: str,
               expected_columns: Optional[List[str]] = None) -> Tuple[pd.DataFrame, int]:
    """Read csv / xls / xlsx into a DataFrame.

    When ``expected_columns`` (the category's canonical schema) is given, the table is
    located by NAME: the header row is the one matching the most canonical names, and only
    the columns spanning those matched headers are kept. Anything sitting elsewhere on the
    sheet - a toggle/parameter cell like ``$AB$2``, notes, a totals block - is left out, so
    it can't be mistaken for an extra column and fail the schema check.

    Without ``expected_columns`` it falls back to generic header auto-detection.
    """
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        raw = pd.read_csv(io.BytesIO(file_bytes), header=None, dtype=object,
                          keep_default_na=False, skip_blank_lines=False)
        rows = raw.values.tolist()
    else:
        engine = "xlrd" if name.endswith(".xls") else "openpyxl"
        raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=object, engine=engine)
        rows = raw.values.tolist()
    if not rows:
        raise ValueError(f"{filename}: no rows found")
    rows = [[None if pd.isna(c) else c for c in r] for r in rows]

    if expected_columns:
        want = {_nkey(c) for c in expected_columns}
        scored = [
            (sum(1 for c in row if _nkey(c) in want), -i, i)
            for i, row in enumerate(rows[:_SCAN_ROWS])
        ]
        hits, _, hr = max(scored)
        if hits > 0:
            matched = [j for j, c in enumerate(rows[hr]) if _nkey(c) in want]
            lo, hi = min(matched), max(matched)
            headers = _headers_from_row([c for c in rows[hr][lo:hi + 1]])
            # the table is the contiguous block under the header - stop at the first
            # blank row (which normally separates it from a totals/notes block below).
            body = []
            for r in rows[hr + 1:]:
                cells = r[lo:hi + 1]
                if not any(_norm(c) for c in cells):
                    break
                body.append(cells)
            df = pd.DataFrame(body, columns=headers)
            return df.reset_index(drop=True), hr
        # no canonical header found: return the generic read so validate_schema can
        # report a clear "missing / unexpected" message rather than a crash.

    hr = detect_header_row(rows)
    headers = _headers_from_row(rows[hr])
    df = pd.DataFrame(rows[hr + 1:], columns=headers).dropna(axis=0, how="all").reset_index(drop=True)
    df = df.loc[:, [not (h.startswith("Column") and df[h].isna().all()) for h in df.columns]]
    return df, hr


def validate_schema(columns: List[str], canonical_schema: List[str]) -> Dict[str, Any]:
    """Exact match: same columns, same order, nothing extra (case/space-insensitive)."""
    def key(s: str) -> str:
        return re.sub(r"\s+", " ", str(s)).strip().lower()

    got = [key(c) for c in columns]
    want = [key(c) for c in canonical_schema]
    got_set, want_set = set(got), set(want)
    missing = [c for c in canonical_schema if key(c) not in got_set]
    extra = [c for c in columns if key(c) not in want_set]
    order_ok = got == want
    ok = not missing and not extra and order_ok
    if ok:
        msg = "Schema matches."
    else:
        parts = []
        if missing:
            parts.append("missing: " + ", ".join(missing))
        if extra:
            parts.append("unexpected: " + ", ".join(extra))
        if not missing and not extra and not order_ok:
            parts.append("column order differs from the sample")
        msg = "; ".join(parts)
    return {"ok": ok, "missing": missing, "extra": extra,
            "order_ok": order_ok, "message": msg}


# ------------------------------------------------------------------------- lookups
class Lookups:
    """Access to the category's lookup tables (persisted from the sample's mapping sheets).

    ``tables`` maps sheet name -> ``{"headers": [...], "records": [ {..}, .. ],
    "header_row": int}``.
    """

    def __init__(self, tables: Dict[str, Dict[str, Any]]):
        self._t = tables or {}
        self._df_cache: Dict[str, pd.DataFrame] = {}
        self._grid_cache: Dict[str, List[List[Any]]] = {}

    def has(self, name: str) -> bool:
        return name in self._t

    def df(self, name: str) -> pd.DataFrame:
        if name not in self._df_cache:
            spec = self._t.get(name)
            if not spec:
                raise KeyError(f"lookup sheet '{name}' is not defined for this category")
            self._df_cache[name] = pd.DataFrame(spec["records"], columns=spec["headers"])
        return self._df_cache[name]

    def grid(self, name: str) -> List[List[Any]]:
        if name not in self._grid_cache:
            spec = self._t[name]
            hr = int(spec.get("header_row", 0))
            grid: List[List[Any]] = [[] for _ in range(hr)]
            grid.append(list(spec["headers"]))
            for rec in spec["records"]:
                grid.append([rec.get(h) for h in spec["headers"]])
            self._grid_cache[name] = grid
        return self._grid_cache[name]

    def cell(self, name: str, a1: str) -> Any:
        m = _A1_RE.match(a1.strip())
        if not m:
            raise ValueError(f"bad cell reference {a1!r}")
        col = col_letter_to_index(m.group(1))
        row = int(m.group(2)) - 1
        grid = self.grid(name)
        if row < len(grid) and col < len(grid[row]):
            return grid[row][col]
        return np.nan

    def vlookup(self, key, name: str, col_index: int, approximate: bool = False):
        frame = self.df(name)
        if frame.shape[1] < col_index:
            raise ValueError(f"VLOOKUP col index {col_index} out of range for '{name}'")
        keys = frame.iloc[:, 0]
        vals = frame.iloc[:, int(col_index) - 1]
        mapping = dict(zip(keys, vals))
        if isinstance(key, pd.Series):
            if approximate:
                sorted_keys = np.sort(keys.values)
                idx = np.searchsorted(sorted_keys, key.values, side="right") - 1
                idx = np.clip(idx, 0, len(sorted_keys) - 1)
                approx_keys = sorted_keys[idx]
                return pd.Series([mapping.get(k, np.nan) for k in approx_keys],
                                 index=key.index)
            return key.map(mapping)
        return mapping.get(key, np.nan)


# ---------------------------------------------------------------- eval helper funcs
def _series_index(*args):
    for a in args:
        if isinstance(a, pd.Series):
            return a.index
    return None


def _as_frame(args, index):
    cols = []
    for a in args:
        cols.append(a if isinstance(a, pd.Series) else pd.Series([a] * len(index), index=index))
    return pd.concat(cols, axis=1)


def IF(cond, a, b):
    idx = _series_index(cond, a, b)
    if idx is not None:
        return pd.Series(np.where(cond, a, b), index=idx)
    return a if cond else b


def IFERROR(value, fallback):
    if isinstance(value, pd.Series):
        return value.where(value.notna() & ~value.isin([np.inf, -np.inf]), fallback)
    try:
        if value is None or (isinstance(value, float) and (np.isnan(value) or np.isinf(value))):
            return fallback
        return value
    except TypeError:
        return value


def _reduce_logical(op, args, empty):
    idx = _series_index(*args)
    if idx is None:
        res = empty
        for a in args:
            res = op(res, bool(a))
        return res
    frame = _as_frame(args, idx).astype(bool)
    out = frame.iloc[:, 0]
    for c in range(1, frame.shape[1]):
        out = op(out, frame.iloc[:, c])
    return out


def AND(*args):
    return _reduce_logical(np.logical_and, args, True)


def OR(*args):
    return _reduce_logical(np.logical_or, args, False)


def NOT(x):
    return ~x if isinstance(x, pd.Series) else (not x)


def _rowwise(func, args):
    idx = _series_index(*args)
    if idx is None:
        return func([a for a in args])
    return _as_frame(args, idx).apply(lambda r: func(list(r.values)), axis=1)


def MIN(*a):
    return _rowwise(lambda xs: np.nanmin(pd.to_numeric(pd.Series(xs), errors="coerce")), a)


def MAX(*a):
    return _rowwise(lambda xs: np.nanmax(pd.to_numeric(pd.Series(xs), errors="coerce")), a)


def SUM(*a):
    return _rowwise(lambda xs: np.nansum(pd.to_numeric(pd.Series(xs), errors="coerce")), a)


def AVERAGE(*a):
    return _rowwise(lambda xs: np.nanmean(pd.to_numeric(pd.Series(xs), errors="coerce")), a)


def COUNT(*a):
    return _rowwise(lambda xs: pd.to_numeric(pd.Series(xs), errors="coerce").notna().sum(), a)


def ROUND(x, n=0):
    return np.round(pd.to_numeric(x, errors="coerce") if isinstance(x, pd.Series) else x, int(n))


def ROUNDUP(x, n=0):
    f = 10 ** int(n)
    return np.ceil((x if not isinstance(x, pd.Series) else pd.to_numeric(x, errors="coerce")) * f) / f


def ROUNDDOWN(x, n=0):
    f = 10 ** int(n)
    return np.floor((x if not isinstance(x, pd.Series) else pd.to_numeric(x, errors="coerce")) * f) / f


def ABS(x):
    return x.abs() if isinstance(x, pd.Series) else abs(x)


def INT(x):
    return np.floor(pd.to_numeric(x, errors="coerce")) if isinstance(x, pd.Series) else int(x)


def MOD(a, b):
    return a % b


def SQRT(x):
    return np.sqrt(x)


def POWER(x, p):
    return x ** p


def _str_series(x):
    return x.astype("string") if isinstance(x, pd.Series) else pd.Series([str(x)])


def CONCAT(*args):
    idx = _series_index(*args)
    if idx is None:
        return "".join("" if a is None else str(a) for a in args)
    frame = _as_frame(args, idx)
    return frame.apply(lambda r: "".join("" if v is None else str(v) for v in r.values), axis=1)


def LEFT(x, n=1):
    return x.astype("string").str[: int(n)] if isinstance(x, pd.Series) else str(x)[: int(n)]


def RIGHT(x, n=1):
    return x.astype("string").str[-int(n):] if isinstance(x, pd.Series) else str(x)[-int(n):]


def MID(x, start, length):
    s = int(start) - 1
    if isinstance(x, pd.Series):
        return x.astype("string").str[s: s + int(length)]
    return str(x)[s: s + int(length)]


def LEN(x):
    return x.astype("string").str.len() if isinstance(x, pd.Series) else len(str(x))


def UPPER(x):
    return x.astype("string").str.upper() if isinstance(x, pd.Series) else str(x).upper()


def LOWER(x):
    return x.astype("string").str.lower() if isinstance(x, pd.Series) else str(x).lower()


def TRIM(x):
    return x.astype("string").str.strip() if isinstance(x, pd.Series) else str(x).strip()


def TEXT(x, _fmt=None):
    return x.astype("string") if isinstance(x, pd.Series) else str(x)


def VALUE(x):
    return pd.to_numeric(x, errors="coerce")


def _dt(x):
    return pd.to_datetime(x, errors="coerce")


def YEAR(x):
    return _dt(x).dt.year if isinstance(x, pd.Series) else _dt(x).year


def MONTH(x):
    return _dt(x).dt.month if isinstance(x, pd.Series) else _dt(x).month


def DAY(x):
    return _dt(x).dt.day if isinstance(x, pd.Series) else _dt(x).day


def TODAY():
    return pd.Timestamp.today().normalize()


def NOW():
    return pd.Timestamp.now()


def ISBLANK(x):
    return x.isna() | (x.astype("string").str.len() == 0) if isinstance(x, pd.Series) else (x is None or x == "")


def ISNUMBER(x):
    return pd.to_numeric(x, errors="coerce").notna() if isinstance(x, pd.Series) else isinstance(x, (int, float))


def ISERROR(x):
    return x.isna() if isinstance(x, pd.Series) else (x is None)


# ------------------------------------------------------------------------- evaluate
class ToggleValue(str):
    """A toggle's value, usable both ways a formula might use it:

    * as text - ``IF($AB$2="Yes", ...)`` -> ``PARAM('X') == 'Yes'`` (plain string equality,
      since this *is* a ``str``), and
    * as a boolean - ``IF(IncludeTax, ...)`` -> ``IF(PARAM('X'), ...)`` (truthy only for
      Yes/True/Y/1, via the overridden ``__bool__``).
    """

    def __bool__(self) -> bool:
        return self.strip().lower() in ("yes", "true", "y", "1")


def _toggle_value(raw: Any) -> ToggleValue:
    return ToggleValue("" if raw is None else str(raw))


def evaluate(df: pd.DataFrame, recipe: Dict[str, Any],
             lookups: Optional[Lookups] = None) -> pd.DataFrame:
    """Return a copy of ``df`` with the recipe's computed columns appended (values)."""
    lookups = lookups or Lookups({})
    toggles = {t["name"]: _toggle_value(t.get("value")) for t in recipe.get("toggles", [])}
    work = df.copy()
    # numeric coercion for canonical columns keeps arithmetic sane; keep original too.
    ns = {
        "np": np, "pd": pd, "df": work,
        "IF": IF, "IFERROR": IFERROR, "AND": AND, "OR": OR, "NOT": NOT,
        "MIN": MIN, "MAX": MAX, "SUM": SUM, "AVERAGE": AVERAGE, "COUNT": COUNT,
        "ROUND": ROUND, "ROUNDUP": ROUNDUP, "ROUNDDOWN": ROUNDDOWN, "ABS": ABS,
        "INT": INT, "MOD": MOD, "SQRT": SQRT, "POWER": POWER,
        "CONCAT": CONCAT, "LEFT": LEFT, "RIGHT": RIGHT, "MID": MID, "LEN": LEN,
        "UPPER": UPPER, "LOWER": LOWER, "TRIM": TRIM, "TEXT": TEXT, "VALUE": VALUE,
        "YEAR": YEAR, "MONTH": MONTH, "DAY": DAY, "TODAY": TODAY, "NOW": NOW,
        "ISBLANK": ISBLANK, "ISNUMBER": ISNUMBER, "ISERROR": ISERROR,
        "CELL": lookups.cell,
        "VLOOKUP": lambda key, sheet, col_index, approx=False: lookups.vlookup(
            key, sheet, col_index, bool(approx)),
        "LOOKUP_TABLE": lookups.df,
        "PARAM": lambda name: toggles.get(name, toggles.get(str(name), np.nan)),
        "MANUAL": lambda fn: (_ for _ in ()).throw(
            ValueError(f"formula uses {fn} which needs a manual pandas expression")),
    }
    errors: List[str] = []
    for col in sorted(recipe.get("computed_columns", []), key=lambda c: c.get("position", 1e9)):
        expr = col.get("pandas_expr") or "np.nan"
        try:
            result = eval(expr, {"__builtins__": {}}, ns)  # noqa: S307 - trusted admin config
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{col['name']}: {exc}")
            result = np.nan
        work[col["name"]] = result
        ns["df"] = work
    if errors:
        work.attrs["compute_errors"] = errors
    return work


# -------------------------------------------------------------------- output workbook
def ordered_columns(recipe: Dict[str, Any]) -> List[str]:
    canonical = list(recipe.get("canonical_schema", []))
    computed = sorted(recipe.get("computed_columns", []), key=lambda c: c.get("position", 1e9))
    total = len(canonical) + len(computed)
    slots: List[Optional[str]] = [None] * total
    for c in computed:
        pos = c.get("position")
        if isinstance(pos, int) and 0 <= pos < total and slots[pos] is None:
            slots[pos] = c["name"]
    it = iter(canonical)
    for i in range(total):
        if slots[i] is None:
            slots[i] = next(it, None)
    # any leftovers (bad positions) appended
    leftover = [c["name"] for c in computed if c["name"] not in slots] + [c for c in it]
    return [s for s in slots if s] + leftover


def build_workbook(values_df: pd.DataFrame, recipe: Dict[str, Any],
                   lookups: Optional[Lookups] = None,
                   data_sheet_name: str = "Data") -> bytes:
    """Build the output ``.xlsx``: computed columns carry the live Excel formula."""
    lookups = lookups or Lookups({})
    cols = [c for c in ordered_columns(recipe) if c in values_df.columns] or list(values_df.columns)
    computed_by_name = {c["name"]: c for c in recipe.get("computed_columns", [])}
    # a computed formula's {tok:Name} placeholders (see sample_parser's fixed-cell
    # detection) point at whichever row that toggle ends up on in the Parameters sheet.
    toggle_row = {t["name"]: i + 2 for i, t in enumerate(recipe.get("toggles", []))}

    def _resolve_tokens(formula: str) -> str:
        return _TOK_RE.sub(
            lambda m: f"Parameters!$B${toggle_row[m.group(1)]}"
            if m.group(1) in toggle_row else "FALSE()",
            formula,
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = data_sheet_name[:31]
    ws.append(cols)
    for r_off, (_, row) in enumerate(values_df.iterrows()):
        excel_row = r_off + 2
        out_row = []
        for c in cols:
            if c in computed_by_name:
                formula = computed_by_name[c].get("excel_formula", "")
                if formula:
                    formula = _resolve_tokens(formula.replace("{r}", str(excel_row)))
                out_row.append(formula if formula else row.get(c))
            else:
                val = row.get(c)
                out_row.append(None if (isinstance(val, float) and np.isnan(val)) else val)
        ws.append(out_row)

    # lookup sheets
    for name, spec in getattr(lookups, "_t", {}).items():
        sh = wb.create_sheet(title=str(name)[:31])
        for _ in range(int(spec.get("header_row", 0))):
            sh.append([])
        sh.append(list(spec["headers"]))
        for rec in spec["records"]:
            sh.append([rec.get(h) for h in spec["headers"]])

    # parameters
    toggles = recipe.get("toggles", [])
    if toggles:
        ps = wb.create_sheet(title="Parameters")
        ps.append(["Name", "Value"])
        for t in toggles:
            ps.append([t["name"], t.get("value")])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def lookups_from_sample(analysis: Dict[str, Any], file_bytes: bytes) -> Dict[str, Dict[str, Any]]:
    """Build the persisted lookup-table spec dict from a sample analysis + the file bytes."""
    from .sample_parser import sheet_to_records
    tables: Dict[str, Dict[str, Any]] = {}
    for lk in analysis.get("lookups", []):
        headers, records = sheet_to_records(file_bytes, lk["sheet"])
        tables[lk["sheet"]] = {
            "headers": headers, "records": records,
            "header_row": lk.get("header_row", 0),
        }
    return tables
