"""Row-level comparison of configured column pairs across every computed output file.

A comparison rule (configured in the Admin UI) is::

    { "left": ["<col>", ...], "right": ["<col>", ...],
      "type": "numeric" | "text", "tolerance": 0.01 }

``left`` / ``right`` may be a single column or a list; columns are paired by position
(1st↔1st, 2nd↔2nd …) and a single column on one side is broadcast against several on the
other.

``build_comparison_workbook`` returns one ``.xlsx``:
  * **Summary**   – per file: rows, matched rows, mismatched rows, checks, diffs
  * **Matched**   – one row per column-pair that agreed, WITH the actual values
  * **Mismatched**– one row per column-pair that differed, WITH the actual values
  * **Issues**    – compared columns that came out entirely empty for a file
  * **_meta**     – generation time + rule list
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import openpyxl

DETAIL_COLUMNS = ["source_file", "row", "left_col", "left_val",
                  "right_col", "right_val", "delta", "status"]


def _num(x):
    return pd.to_numeric(pd.Series([x]), errors="coerce").iloc[0]


def _clean(v):
    if v is None:
        return None
    if isinstance(v, float) and np.isnan(v):
        return None
    return v.item() if hasattr(v, "item") else v


def _cmp_cell(left: Any, right: Any, rule: Dict[str, Any]) -> Tuple[bool, float]:
    rtype = rule.get("type", "numeric")
    if rtype == "numeric":
        lv, rv = _num(left), _num(right)
        if pd.isna(lv) and pd.isna(rv):
            return True, 0.0
        if pd.isna(lv) or pd.isna(rv):
            return False, float("nan")
        delta = abs(float(lv) - float(rv))
        return delta <= float(rule.get("tolerance", 0) or 0), delta
    ls = "" if left is None or (isinstance(left, float) and np.isnan(left)) else str(left).strip()
    rs = "" if right is None or (isinstance(right, float) and np.isnan(right)) else str(right).strip()
    if not rule.get("case_sensitive"):
        ls, rs = ls.lower(), rs.lower()
    return ls == rs, 0.0 if ls == rs else float("nan")


def _as_list(v) -> List[str]:
    if v is None:
        return []
    return [str(x) for x in v] if isinstance(v, (list, tuple)) else [str(v)]


def rule_pairs(rule: Dict[str, Any]) -> List[Tuple[str, str]]:
    lefts, rights = _as_list(rule.get("left")), _as_list(rule.get("right"))
    if len(lefts) == 1 and len(rights) > 1:
        lefts = lefts * len(rights)
    if len(rights) == 1 and len(lefts) > 1:
        rights = rights * len(lefts)
    return [(l, r) for l, r in zip(lefts, rights) if l and r]


def _all_pairs(rules: List[Dict[str, Any]]) -> List[Tuple[str, str, Dict[str, Any]]]:
    return [(l, r, rule) for rule in rules for l, r in rule_pairs(rule)]


def compare_file(filename: str, values_df: pd.DataFrame,
                 rules: List[Dict[str, Any]],
                 key_columns: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    """Return one detail dict per (row, column-pair). ``status`` is
    ``match`` / ``diff`` / ``column missing``."""
    key_columns = [k for k in (key_columns or []) if k in values_df.columns]
    pairs = _all_pairs(rules)
    detail: List[Dict[str, Any]] = []

    for pos, (_, row) in enumerate(values_df.iterrows()):
        excel_row = pos + 2
        key_part = {f"key:{k}": _clean(row.get(k)) for k in key_columns}
        for lcol, rcol, rule in pairs:
            rec = {"source_file": filename, "row": excel_row, **key_part,
                   "left_col": lcol, "right_col": rcol}
            if lcol not in values_df.columns or rcol not in values_df.columns:
                missing = [c for c in (lcol, rcol) if c not in values_df.columns]
                detail.append({**rec, "left_val": None, "right_val": None, "delta": None,
                               "status": "column missing: " + ", ".join(missing)})
                continue
            ok, delta = _cmp_cell(row.get(lcol), row.get(rcol), rule)
            detail.append({
                **rec,
                "left_val": _clean(row.get(lcol)),
                "right_val": _clean(row.get(rcol)),
                "delta": None if (isinstance(delta, float) and np.isnan(delta)) else delta,
                "status": "match" if ok else "diff",
            })
    return detail


def _empty_column_issues(filename: str, values_df: pd.DataFrame,
                         rules: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Flag compared columns that exist but are entirely blank for this file - the usual
    sign that a computed column's Python expression evaluated to nothing."""
    issues = []
    cols = {c for l, r, _ in _all_pairs(rules) for c in (l, r)}
    for c in sorted(cols):
        if c in values_df.columns and len(values_df) and values_df[c].isna().all():
            issues.append({"source_file": filename, "column": c,
                           "note": "column is present but every value is empty - "
                                   "check this column's Python expression in Admin"})
    return issues


def run_comparison(files: List[Tuple[str, pd.DataFrame]], recipe: Dict[str, Any]
                   ) -> Dict[str, pd.DataFrame]:
    rules = recipe.get("comparison", [])
    key_columns = recipe.get("comparison_keys", [])
    key_cols_present = [f"key:{k}" for k in key_columns]

    detail_rows: List[Dict[str, Any]] = []
    summary: List[Dict[str, Any]] = []
    issues: List[Dict[str, Any]] = []

    for filename, vdf in files:
        d = compare_file(filename, vdf, rules, key_columns)
        detail_rows.extend(d)
        issues.extend(_empty_column_issues(filename, vdf, rules))

        by_row: Dict[Any, List[str]] = {}
        for rec in d:
            by_row.setdefault(rec["row"], []).append(rec["status"])
        matched_rows = sum(1 for st in by_row.values() if all(s == "match" for s in st))
        summary.append({
            "source_file": filename,
            "rows": len(vdf),
            "matched_rows": matched_rows,
            "mismatched_rows": len(by_row) - matched_rows,
            "checks": len(d),
            "diffs": sum(1 for rec in d if rec["status"] != "match"),
        })

    cols = ([c for c in DETAIL_COLUMNS[:2]] + key_cols_present + DETAIL_COLUMNS[2:])
    detail = pd.DataFrame(detail_rows)
    if not detail.empty:
        detail = detail.reindex(columns=[c for c in cols if c in detail.columns]
                                + [c for c in detail.columns if c not in cols])

    matched = detail[detail["status"] == "match"] if not detail.empty else detail
    mismatched = detail[detail["status"] != "match"] if not detail.empty else detail

    return {
        "Summary": pd.DataFrame(summary),
        "Matched": matched.reset_index(drop=True),
        "Mismatched": mismatched.reset_index(drop=True),
        "Issues": pd.DataFrame(issues),
    }


def build_comparison_workbook(files: List[Tuple[str, pd.DataFrame]],
                              recipe: Dict[str, Any]) -> bytes:
    sheets = run_comparison(files, recipe)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in ("Summary", "Matched", "Mismatched", "Issues"):
        frame = sheets.get(name, pd.DataFrame())
        ws = wb.create_sheet(title=name)
        if frame is None or frame.empty:
            ws.append([f"(no {name.lower()} rows)"])
            continue
        ws.append(list(frame.columns))
        for _, row in frame.iterrows():
            ws.append([
                None if (isinstance(v, float) and np.isnan(v))
                else (v.item() if hasattr(v, "item") else v)
                for v in row.values
            ])
    meta = wb.create_sheet(title="_meta")
    meta.append(["generated", datetime.utcnow().isoformat() + "Z"])
    meta.append(["category", recipe.get("name", recipe.get("id", ""))])
    meta.append(["rules", "; ".join(
        f"{l} vs {r}" for rule in recipe.get("comparison", []) for l, r in rule_pairs(rule)
    )])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
